import mysql.connector
import re
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import logging
import os

# ======================================================================
# 🔧 CONFIGURATION SECTION - EDIT THESE VALUES AS NEEDED
# ======================================================================

CONFIG = {
    # Database Connection Settings
    'DATABASE': {
        'host': '127.0.0.1',
        'database': 'nests_emails',
        'table': 'email_lists_2',
        'user': 'root',
        'password': '',
        'charset': 'utf8mb4',
        'collation': 'utf8mb4_unicode_ci'
    },
    
    # Application Settings
    'APPLICATION': {
        'hostel_name': 'Aguere',
        'consent_default': '1',  # Default consent value (1=true, 0=false)
        'excel_filename': 'guests.xlsx',
        'date_format': '%d/%m/%Y'  # Input date format from Excel
    },
    
    # Email Filtering Settings
    'EMAIL_FILTERING': {
        'fake_domains': [
            '@guest.booking.com',
            '@expediapartnercentral.com', 
            '@noemail.com',
            '@airbnb.com'
        ]
    },
    
    # Logging Settings
    'LOGGING': {
        'log_directory': 'logs',
        'log_level': logging.INFO,
        'show_results_limit': 10  # Number of records to show in final summary
    },
    
    # Excel Column Mapping (0-based index)
    'EXCEL_COLUMNS': {
        'first_name': 0,      # Nombre
        'last_name': 1,       # Apellido
        'email': 2,           # Correo electrónico
        'phone': 3,           # Teléfono
        'city': 6,            # Ciudad
        'country': 7,         # País
        'postal_code': 9,     # Código postal
        'nights': 11,         # Noches de estadía
        'last_stay': 13       # Última estadía
    }
}

# ======================================================================
# 📋 END OF CONFIGURATION SECTION
# ======================================================================

def setup_logging():
    """
    Setup logging to both console and file with timestamp
    """
    # Create logs directory if it doesn't exist
    log_dir = CONFIG['LOGGING']['log_directory']
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    
    # Create log filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"{log_dir}/guest_extraction_{timestamp}.log"
    
    # Configure logging
    logging.basicConfig(
        level=CONFIG['LOGGING']['log_level'],
        format='%(asctime)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()  # This keeps console output
        ]
    )
    
    return log_filename

def log_and_print(message):
    """
    Print to console and log to file
    """
    print(message)
    logging.info(message)

def is_valid_email(email):
    """
    Check if email format is valid
    Returns True if email has correct format, False otherwise
    """
    if not email or not isinstance(email, str):
        return False
    
    # Email pattern: something@domain.extension
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email.strip()) is not None

def is_not_booking_email(email):
    """
    Check if email is NOT from booking platforms or fake email services
    Returns True if email is real, False if it's from booking/fake services
    """
    if not email or not isinstance(email, str):
        return False
    
    email_lower = email.lower().strip()
    
    # Get fake domains from configuration
    fake_domains = CONFIG['EMAIL_FILTERING']['fake_domains']
    
    # Check if email ends with any of the fake domains
    for domain in fake_domains:
        if email_lower.endswith(domain):
            return False
    
    return True

def parse_date(date_string):
    """
    Parse date from configured format to YYYY-MM-DD
    Returns formatted date string or None if invalid
    """
    if not date_string or not isinstance(date_string, str):
        return None
    
    try:
        # Parse date using configured format
        date_format = CONFIG['APPLICATION']['date_format']
        date_obj = datetime.strptime(date_string.strip(), date_format)
        return date_obj.strftime("%Y-%m-%d")
    except ValueError:
        log_and_print(f"⚠️  Invalid date format: {date_string} (expected format: {CONFIG['APPLICATION']['date_format']})")
        return None

def calculate_checkin_date(checkout_date_str, nights):
    """
    Calculate checkin date by subtracting nights from checkout date
    Returns checkin date string in YYYY-MM-DD format or None
    """
    if not checkout_date_str or not nights:
        return None
    
    try:
        # Parse checkout date
        checkout_date = datetime.strptime(checkout_date_str, "%Y-%m-%d")
        # Subtract nights to get checkin date
        checkin_date = checkout_date - timedelta(days=int(nights))
        return checkin_date.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        log_and_print(f"⚠️  Could not calculate checkin date from {checkout_date_str} and {nights} nights")
        return None

def connect_to_database():
    """
    Connect to MySQL database using configuration settings
    Returns database connection
    """
    try:
        db_config = CONFIG['DATABASE']
        conn = mysql.connector.connect(
            host=db_config['host'],
            database=db_config['database'],
            user=db_config['user'],
            password=db_config['password'],
            charset=db_config['charset'],
            collation=db_config['collation']
        )
        log_and_print(f"✅ Connected to MySQL database '{db_config['database']}' successfully!")
        return conn
    except mysql.connector.Error as e:
        log_and_print(f"❌ Error connecting to MySQL database: {e}")
        return None

def read_excel_file(filename):
    """
    Read the Excel file and extract guest data using configured column mappings
    Returns list of tuples with all guest information
    """
    try:
        # Load the Excel file
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        guests_data = []
        col_map = CONFIG['EXCEL_COLUMNS']
        
        # Skip header row (row 1), start from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Extract data using configured column mappings
            first_name = str(row[col_map['first_name']]).strip() if row[col_map['first_name']] else ""
            last_name = str(row[col_map['last_name']]).strip() if row[col_map['last_name']] else ""
            email = str(row[col_map['email']]).strip() if row[col_map['email']] else ""
            phone = str(row[col_map['phone']]).strip() if row[col_map['phone']] else ""
            city = str(row[col_map['city']]).strip() if row[col_map['city']] else ""
            country = str(row[col_map['country']]).strip() if row[col_map['country']] else ""
            postal_code = str(row[col_map['postal_code']]).strip() if row[col_map['postal_code']] else ""
            nights = row[col_map['nights']] if row[col_map['nights']] else None
            last_stay = str(row[col_map['last_stay']]).strip() if row[col_map['last_stay']] else ""
            
            # Save ALL records (even without complete data)
            if first_name or last_name or email:  # At least one identifier
                guests_data.append({
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'phone': phone,
                    'city': city,
                    'country': country,
                    'postal_code': postal_code,
                    'nights': nights,
                    'last_stay': last_stay,
                    'row_number': row_num
                })
        
        log_and_print(f"📖 Read {len(guests_data)} guest records from Excel file '{filename}'")
        return guests_data
        
    except FileNotFoundError:
        log_and_print(f"❌ Error: File '{filename}' not found!")
        return []
    except Exception as e:
        log_and_print(f"❌ Error reading Excel file: {e}")
        return []

def process_and_save_guests(conn, guests_data):
    """
    Process guest data and save ALL entries to database (no filtering)
    Saves users even without dates or with incomplete data
    """
    cursor = conn.cursor()
    
    total_processed = 0
    saved_count = 0
    booking_emails = 0
    invalid_emails = 0
    date_issues = 0
    no_dates_saved = 0
    duplicates_skipped = 0
    other_errors = 0
    
    log_and_print("\n🔄 Processing ALL guests (no filtering, saving even without dates)...")
    log_and_print("-" * 80)
    
    for guest in guests_data:
        total_processed += 1
        
        # Check email quality for reporting (but don't filter)
        is_valid_email_format = is_valid_email(guest['email'])
        is_booking_email = not is_not_booking_email(guest['email'])
        
        if not is_valid_email_format:
            invalid_emails += 1
        
        if is_booking_email:
            booking_emails += 1
        
        # Process dates (but save even if dates are missing)
        checkout_date = parse_date(guest['last_stay']) if guest['last_stay'] else None
        checkin_date = None
        
        if checkout_date and guest['nights']:
            checkin_date = calculate_checkin_date(checkout_date, guest['nights'])
        
        if not checkout_date:
            date_issues += 1
            no_dates_saved += 1
        
        # Save ALL records to database (no filtering, even without dates)
        try:
            table_name = CONFIG['DATABASE']['table']
            insert_query = f"""
                INSERT INTO {table_name} 
                (first_name, last_name, email, phone, checkin, checkout, country, city, postal_code, consent, hostel, created_at, updated_at) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            """
            
            values = (
                guest['first_name'] if guest['first_name'] else None,
                guest['last_name'] if guest['last_name'] else None,
                guest['email'] if guest['email'] else None,
                guest['phone'] if guest['phone'] else None,
                checkin_date,  # Can be None
                checkout_date,  # Can be None
                guest['country'] if guest['country'] else None,
                guest['city'] if guest['city'] else None,
                guest['postal_code'] if guest['postal_code'] != 'None' and guest['postal_code'] else None,
                CONFIG['APPLICATION']['consent_default'],  # consent from config
                CONFIG['APPLICATION']['hostel_name']  # hostel name from config
            )
            
            cursor.execute(insert_query, values)
            saved_count += 1
            
            # Create status indicators for reporting
            status_flags = []
            if not is_valid_email_format:
                status_flags.append("📧❌")
            if is_booking_email:
                status_flags.append("🏨")
            if not checkout_date:
                status_flags.append("📅❌")
            
            status_text = " ".join(status_flags) if status_flags else "✅"
            
            # Show dates info
            date_info = ""
            if checkin_date and checkout_date:
                date_info = f" | {checkin_date} to {checkout_date}"
            elif checkout_date:
                date_info = f" | checkout: {checkout_date}"
            else:
                date_info = " | no dates (SAVED ANYWAY)"
            
            log_and_print(f"{status_text} Row {guest['row_number']}: {guest['first_name']} {guest['last_name']} - {guest['email']}{date_info}")
            
        except mysql.connector.IntegrityError as e:
            if "Duplicate entry" in str(e):
                duplicates_skipped += 1
                log_and_print(f"⚠️  Row {guest['row_number']}: Duplicate email skipped: {guest['email']}")
            else:
                other_errors += 1
                log_and_print(f"❌ Row {guest['row_number']}: Database integrity error: {e}")
        except Exception as e:
            other_errors += 1
            log_and_print(f"❌ Row {guest['row_number']}: Unexpected error: {e}")
    
    # Save all changes to database
    conn.commit()
    
    # Print summary
    log_and_print("\n" + "=" * 80)
    log_and_print("📊 PROCESSING SUMMARY (ALL RECORDS SAVED):")
    log_and_print("=" * 80)
    log_and_print(f"Total records processed: {total_processed}")
    log_and_print(f"✅ Records saved to database: {saved_count}")
    log_and_print(f"📅❌ Records saved WITHOUT dates: {no_dates_saved}")
    log_and_print(f"⚠️  Duplicate emails skipped: {duplicates_skipped}")
    log_and_print(f"❌ Other errors: {other_errors}")
    log_and_print("")
    log_and_print("📊 DATA QUALITY REPORT:")
    log_and_print(f"📧❌ Invalid email formats: {invalid_emails}")
    log_and_print(f"🏨 Booking/Platform emails: {booking_emails}")
    log_and_print(f"📅❌ Date parsing issues: {date_issues}")
    log_and_print("=" * 80)
    log_and_print("Legend: 📧❌=Invalid Email | 🏨=Booking/Platform | 📅❌=Date Issue | ✅=Clean Record")

def show_database_contents(conn, limit=None):
    """
    Display some saved records from the database
    """
    if limit is None:
        limit = CONFIG['LOGGING']['show_results_limit']
    
    cursor = conn.cursor()
    table_name = CONFIG['DATABASE']['table']
    cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
    total = cursor.fetchone()[0]
    
    log_and_print(f"\n📊 Database now contains {total} guest records in '{table_name}' table")
    
    if total > 0:
        log_and_print(f"\nFirst {min(limit, total)} records:")
        log_and_print("-" * 100)
        cursor.execute(f"""
            SELECT first_name, last_name, email, city, country, checkin, checkout 
            FROM {table_name} 
            ORDER BY id DESC 
            LIMIT %s
        """, (limit,))
        
        for row in cursor.fetchall():
            dates = f"{row[5]} to {row[6]}" if row[5] and row[6] else "No dates"
            location = f"{row[3]}, {row[4]}" if row[3] and row[4] else "No location"
            log_and_print(f"{row[0]} {row[1]} - {row[2]} | {location} | {dates}")

def main():
    """
    Main function to run the entire extraction process
    """
    # Setup logging first
    log_filename = setup_logging()
    
    log_and_print(f"🏨 GUEST EMAIL EXTRACTOR FOR {CONFIG['APPLICATION']['hostel_name'].upper()} HOSTEL")
    log_and_print("=" * 60)
    log_and_print(f"📄 Log file: {log_filename}")
    log_and_print(f"🏨 Hostel: {CONFIG['APPLICATION']['hostel_name']}")
    log_and_print(f"📊 Database: {CONFIG['DATABASE']['database']}.{CONFIG['DATABASE']['table']}")
    log_and_print(f"📧 Filtering {len(CONFIG['EMAIL_FILTERING']['fake_domains'])} fake domain types")
    log_and_print("=" * 60)
    
    # Configuration
    excel_filename = CONFIG['APPLICATION']['excel_filename']
    
    # Step 1: Connect to database
    log_and_print("1. Connecting to MySQL database...")
    conn = connect_to_database()
    if not conn:
        log_and_print("❌ Could not connect to database. Exiting.")
        return
    
    # Step 2: Read Excel file
    log_and_print("\n2. Reading Excel file...")
    guests_data = read_excel_file(excel_filename)
    
    if not guests_data:
        log_and_print("❌ No data found. Exiting.")
        conn.close()
        return
    
    # Step 3: Process and save data
    log_and_print("\n3. Processing and saving ALL data (including records without dates)...")
    process_and_save_guests(conn, guests_data)
    
    # Step 4: Show results
    log_and_print("\n4. Showing latest results...")
    show_database_contents(conn)
    
    # Close database connection
    conn.close()
    
    table_name = CONFIG['DATABASE']['table']
    database_name = CONFIG['DATABASE']['database']
    hostel_name = CONFIG['APPLICATION']['hostel_name']
    
    log_and_print(f"\n🎉 Process completed! ALL records saved to '{table_name}' table in '{database_name}' database.")
    log_and_print(f"Note: All records have hostel='{hostel_name}' and consent='{CONFIG['APPLICATION']['consent_default']}'")
    log_and_print("📊 Check the summary above for data quality insights!")
    log_and_print(f"📄 Complete log saved to: {log_filename}")

# Run the script
if __name__ == "__main__":
    main()